from __future__ import annotations

import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg import sql


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_FOLDER = PROJECT_ROOT / "data"
ENV_FILE = PROJECT_ROOT / ".env"


# Excel sheet → PostgreSQL table → expected columns → unique key
TABLES = {
    "PLATFORM": {
        "table": "platform",
        "columns": [
            "platform_id",
            "platform_name",
            "category",
            "ai_native_ide",
            "ide_integrated_copilot",
            "offering_company",
        ],
        "key": ["platform_id"],
    },
    "MODEL_PROVIDER": {
        "table": "model_provider",
        "columns": [
            "provider_id",
            "provider_name",
            "provider_type",
        ],
        "key": ["provider_id"],
    },
    "MODEL_FAMILY": {
        "table": "model_family",
        "columns": [
            "model_family_id",
            "provider_id",
            "family_name",
            "coverage_description",
            "grouping_note",
        ],
        "key": ["model_family_id"],
    },
    "FEATURE": {
        "table": "feature",
        "columns": [
            "feature_id",
            "feature_name",
            "feature_category",
        ],
        "key": ["feature_id"],
    },
    "AVAILABILITY_STATUS": {
        "table": "availability_status",
        "columns": [
            "status_id",
            "status_label",
            "status_group",
            "directly_selectable",
            "counts_as_current",
        ],
        "key": ["status_id"],
    },
    "SUBSCRIPTION_TIER": {
        "table": "subscription_tier",
        "columns": [
            "tier_id",
            "platform_id",
            "tier_name",
            "license_type",
            "usage_credits",
            "rate_limit_window",
            "team_seats",
            "model_access_summary",
            "core_privileges",
            "additional_privileges",
            "not_included",
            "best_suited_for",
        ],
        "key": ["tier_id"],
    },
    "PLAN_PRICING": {
        "table": "plan_pricing",
        "columns": [
            "pricing_id",
            "tier_id",
            "currency",
            "billing_frequency",
            "commitment_months",
            "monthly_list_price",
            "annual_total_price",
            "annual_monthly_equivalent",
            "per_user_monthly",
            "base_team_monthly",
            "minimum_account_charge",
            "pricing_formula_type",
            "effective_date",
        ],
        "key": ["pricing_id"],
    },
    "PLATFORM_MODEL_AVAILABILITY": {
        "table": "platform_model_availability",
        "columns": [
            "platform_id",
            "model_family_id",
            "status_id",
        ],
        "key": ["platform_id", "model_family_id"],
    },
    "PLATFORM_FEATURE_SUPPORT": {
        "table": "platform_feature_support",
        "columns": [
            "platform_id",
            "feature_id",
            "supported",
            "support_value",
            "support_note",
        ],
        "key": ["platform_id", "feature_id"],
    },
}


BOOLEAN_COLUMNS = {
    "ai_native_ide",
    "ide_integrated_copilot",
    "directly_selectable",
    "counts_as_current",
    "supported",
}


def find_workbook() -> Path:
    """Find the Excel workbook inside the data folder."""

    excel_files = [
        path
        for path in DATA_FOLDER.glob("*.xlsx")
        if not path.name.startswith("~$")
    ]

    if len(excel_files) == 0:
        raise FileNotFoundError(
            f"No .xlsx workbook was found in {DATA_FOLDER}"
        )

    if len(excel_files) > 1:
        names = "\n".join(f"- {path.name}" for path in excel_files)
        raise RuntimeError(
            "More than one Excel workbook was found in the data folder.\n"
            "Keep only the workbook you want to import:\n"
            f"{names}"
        )

    return excel_files[0]


def clean_boolean(value: Any) -> bool | None:
    """Convert common Excel Boolean representations."""

    if value is None:
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    if isinstance(value, str):
        cleaned = value.strip().lower()

        if cleaned in {"true", "yes", "1"}:
            return True

        if cleaned in {"false", "no", "0"}:
            return False

    raise ValueError(f"Invalid Boolean value: {value!r}")


def clean_value(
    value: Any,
    column_name: str,
    workbook_epoch: Any,
) -> Any:
    """Prepare an Excel value for PostgreSQL."""

    if isinstance(value, str):
        value = value.strip()

        if value == "":
            return None

    if column_name in BOOLEAN_COLUMNS:
        return clean_boolean(value)

    if column_name == "effective_date":
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            converted = from_excel(value, workbook_epoch)

            if isinstance(converted, datetime):
                return converted.date()

            return converted

        raise ValueError(
            f"Invalid effective_date value: {value!r}"
        )

    return value


def read_excel_data(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read and validate all required Excel sheets."""

    workbook = load_workbook(
        workbook_path,
        data_only=True,
        read_only=True,
    )

    rows_by_sheet: dict[str, list[dict[str, Any]]] = {}

    for sheet_name, settings in TABLES.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Required sheet is missing: {sheet_name}"
            )

        worksheet = workbook[sheet_name]
        expected_columns = settings["columns"]

        actual_headers = [
            cell.value.strip() if isinstance(cell.value, str) else cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]

        # Ignore completely blank cells after the real table.
        actual_headers = [
            header for header in actual_headers if header is not None
        ]

        if actual_headers != expected_columns:
            raise ValueError(
                f"\nHeader mismatch in sheet {sheet_name}.\n"
                f"Expected: {expected_columns}\n"
                f"Found:    {actual_headers}"
            )

        sheet_rows: list[dict[str, Any]] = []

        for excel_row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_col=len(expected_columns),
                values_only=True,
            ),
            start=2,
        ):
            if all(value is None for value in values):
                continue

            cleaned_row = {
                column: clean_value(
                    value,
                    column,
                    workbook.epoch,
                )
                for column, value in zip(expected_columns, values)
            }

            cleaned_row["_excel_row"] = excel_row_number
            sheet_rows.append(cleaned_row)

        rows_by_sheet[sheet_name] = sheet_rows

    return rows_by_sheet


def validate_duplicate_keys(
    rows_by_sheet: dict[str, list[dict[str, Any]]]
) -> None:
    """Check primary and junction-table keys before uploading."""

    for sheet_name, settings in TABLES.items():
        key_columns = settings["key"]
        seen: dict[tuple[Any, ...], int] = {}

        for row in rows_by_sheet[sheet_name]:
            key = tuple(row[column] for column in key_columns)

            if any(value is None for value in key):
                raise ValueError(
                    f"{sheet_name}, Excel row {row['_excel_row']}: "
                    f"missing required key value in {key_columns}"
                )

            if key in seen:
                raise ValueError(
                    f"Duplicate key in {sheet_name}: {key}. "
                    f"Found on Excel rows {seen[key]} "
                    f"and {row['_excel_row']}."
                )

            seen[key] = row["_excel_row"]


def validate_relationships(
    rows_by_sheet: dict[str, list[dict[str, Any]]]
) -> None:
    """Check the main cross-sheet ID relationships."""

    platform_ids = {
        row["platform_id"]
        for row in rows_by_sheet["PLATFORM"]
    }

    provider_ids = {
        row["provider_id"]
        for row in rows_by_sheet["MODEL_PROVIDER"]
    }

    model_family_ids = {
        row["model_family_id"]
        for row in rows_by_sheet["MODEL_FAMILY"]
    }

    feature_ids = {
        row["feature_id"]
        for row in rows_by_sheet["FEATURE"]
    }

    status_ids = {
        row["status_id"]
        for row in rows_by_sheet["AVAILABILITY_STATUS"]
    }

    tier_ids = {
        row["tier_id"]
        for row in rows_by_sheet["SUBSCRIPTION_TIER"]
    }

    for row in rows_by_sheet["MODEL_FAMILY"]:
        if row["provider_id"] not in provider_ids:
            raise ValueError(
                f"MODEL_FAMILY row {row['_excel_row']} refers to "
                f"unknown provider_id {row['provider_id']!r}"
            )

    for row in rows_by_sheet["SUBSCRIPTION_TIER"]:
        if row["platform_id"] not in platform_ids:
            raise ValueError(
                f"SUBSCRIPTION_TIER row {row['_excel_row']} refers to "
                f"unknown platform_id {row['platform_id']!r}"
            )

    for row in rows_by_sheet["PLAN_PRICING"]:
        if row["tier_id"] not in tier_ids:
            raise ValueError(
                f"PLAN_PRICING row {row['_excel_row']} refers to "
                f"unknown tier_id {row['tier_id']!r}"
            )

    for row in rows_by_sheet["PLATFORM_MODEL_AVAILABILITY"]:
        if row["platform_id"] not in platform_ids:
            raise ValueError(
                f"PLATFORM_MODEL_AVAILABILITY row "
                f"{row['_excel_row']} has an unknown platform_id."
            )

        if row["model_family_id"] not in model_family_ids:
            raise ValueError(
                f"PLATFORM_MODEL_AVAILABILITY row "
                f"{row['_excel_row']} has an unknown model_family_id."
            )

        if row["status_id"] not in status_ids:
            raise ValueError(
                f"PLATFORM_MODEL_AVAILABILITY row "
                f"{row['_excel_row']} has an unknown status_id."
            )

    for row in rows_by_sheet["PLATFORM_FEATURE_SUPPORT"]:
        if row["platform_id"] not in platform_ids:
            raise ValueError(
                f"PLATFORM_FEATURE_SUPPORT row "
                f"{row['_excel_row']} has an unknown platform_id."
            )

        if row["feature_id"] not in feature_ids:
            raise ValueError(
                f"PLATFORM_FEATURE_SUPPORT row "
                f"{row['_excel_row']} has an unknown feature_id."
            )

        support_value = row["support_value"]
        supported = row["supported"]

        if support_value is None:
            raise ValueError(
                f"PLATFORM_FEATURE_SUPPORT row "
                f"{row['_excel_row']} has no support_value."
            )

        if not 0 <= float(support_value) <= 1:
            raise ValueError(
                f"PLATFORM_FEATURE_SUPPORT row "
                f"{row['_excel_row']} has support_value outside 0–1."
            )

        if supported != (float(support_value) > 0):
            raise ValueError(
                f"PLATFORM_FEATURE_SUPPORT row "
                f"{row['_excel_row']} has inconsistent "
                f"supported and support_value values."
            )


def insert_sheet(
    cursor: psycopg.Cursor[Any],
    sheet_name: str,
    rows: list[dict[str, Any]],
) -> None:
    """Insert one Excel sheet into its PostgreSQL table."""

    settings = TABLES[sheet_name]
    table_name = settings["table"]
    columns = settings["columns"]

    statement = sql.SQL(
        "INSERT INTO catalog.{} ({}) VALUES ({})"
    ).format(
        sql.Identifier(table_name),
        sql.SQL(", ").join(
            sql.Identifier(column) for column in columns
        ),
        sql.SQL(", ").join(
            sql.Placeholder() for _ in columns
        ),
    )

    values = [
        tuple(row[column] for column in columns)
        for row in rows
    ]

    cursor.executemany(statement, values)


def main() -> None:
    load_dotenv(ENV_FILE)

    database_url = os.getenv("OWNER_DATABASE_URL")

    if not database_url:
        print("ERROR: OWNER_DATABASE_URL was not found in .env")
        sys.exit(1)

    try:
        workbook_path = find_workbook()

        print(f"Workbook found: {workbook_path.name}")
        print("Reading and validating Excel data...")

        rows_by_sheet = read_excel_data(workbook_path)
        validate_duplicate_keys(rows_by_sheet)
        validate_relationships(rows_by_sheet)

        print("Excel validation passed.")
        print("Uploading data to Neon...")

        with psycopg.connect(database_url) as connection:
            with connection.cursor() as cursor:
                # Remove previous imported data safely.
                cursor.execute(
                    """
                    TRUNCATE TABLE
                        catalog.platform_feature_support,
                        catalog.platform_model_availability,
                        catalog.plan_pricing,
                        catalog.subscription_tier,
                        catalog.model_family,
                        catalog.feature,
                        catalog.availability_status,
                        catalog.model_provider,
                        catalog.platform
                    CASCADE;
                    """
                )

                # Parent tables must be inserted before child tables.
                insertion_order = [
                    "PLATFORM",
                    "MODEL_PROVIDER",
                    "FEATURE",
                    "AVAILABILITY_STATUS",
                    "MODEL_FAMILY",
                    "SUBSCRIPTION_TIER",
                    "PLAN_PRICING",
                    "PLATFORM_MODEL_AVAILABILITY",
                    "PLATFORM_FEATURE_SUPPORT",
                ]

                for sheet_name in insertion_order:
                    insert_sheet(
                        cursor,
                        sheet_name,
                        rows_by_sheet[sheet_name],
                    )

        print("\nImport completed successfully!\n")

        for sheet_name in TABLES:
            table_name = TABLES[sheet_name]["table"]
            row_count = len(rows_by_sheet[sheet_name])

            print(f"{table_name}: {row_count} rows")

    except Exception as error:
        print("\nImport failed.")
        print(error)
        sys.exit(1)


if __name__ == "__main__":
    main()